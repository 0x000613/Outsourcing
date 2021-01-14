import os
import bs4
from openpyxl import Workbook
from selenium import webdriver

# 엑셀 데이터 제어 초기설정
write_wb = Workbook()
write_ws = write_wb.active

write_ws["A1"] = "연번"
write_ws["B1"] = "환자"
write_ws["C1"] = "확진일"
write_ws["D1"] = "거주지"
write_ws["E1"] = "여행력"
write_ws["F1"] = "접촉력"
write_ws["G1"] = "조치사항"

# 데이터 폼 초기화
PR = []
DT = []
AR = []
IF = []
CT = []
AC = []

# 현재 작업디렉토리를 BASE_DIR 변수에 저장
try:
    BASE_DIR = os.path.realpath(os.path.dirname("__file__"))
except:
    BASE_DIR = os.path.realpath(os.path.dirname(__file__))
# Selenium 모듈에서 사용할 chrome driver를 로드
try:
    driver = webdriver.Chrome(os.path.join(BASE_DIR, "chromedriver.exe"))
except FileNotFoundError:
    print("크롬 드라이버를 찾을 수 없습니다.")

# 데이타 파서
def getData(pageSource):
    # 새로 로드된 페이지 내용 로드
    PR = pageSource.find_all("td", attrs={"data-tit":"환자 번호"})
    # 확진일
    DT = pageSource.find_all("td", attrs={"data-tit":"확진일"})
    # 거주지
    AR = pageSource.find_all("td", attrs={"data-tit":"거주지"})
    # 여행력
    IF = pageSource.find_all("td", attrs={"data-tit":"여행력"})
    # 접촉력
    CT = pageSource.find_all("td", attrs={"data-tit":"접촉력"})
    # 조치사항
    AC = pageSource.find_all("td", attrs={"data-tit":"조치사항"})
    
    PR.reverse()
    DT.reverse()
    AR.reverse()
    IF.reverse()
    CT.reverse()
    AC.reverse()

    return PR, DT, AR, IF, CT, AC

def main():
    # 크롬 창 크기 설정
    driver.set_window_size(1024, 800)
    # 드라이버 로드 대기
    driver.implicitly_wait(1)
    # 서울 코로나 공공데이터 사이트 접속
    driver.get("http://www.seoul.go.kr/coronaV/coronaStatus.do#status_page_top")

    # 첫번째 페이지로 이동
    firstPageBtn = driver.find_element_by_css_selector('#DataTables_Table_0_paginate > span > a:nth-child(7)')
    driver.execute_script("arguments[0].click();", firstPageBtn)

    # 모든 페이지 크롤링
    while True:
        # 환자 데이터 파싱
        pageSource = bs4.BeautifulSoup(driver.page_source, "html.parser")
        # 바로 위 pageSource를 getData함수에 넣어서 계산된 값을 data 변수에 넣겠다.
        data = getData(pageSource)

        # getData함수를 이용해 return 받은 데이터를 append
        for i in range(len(data[0])):
            PR.append(data[0][i])
            DT.append(data[1][i])
            AR.append(data[2][i])
            IF.append(data[3][i])
            CT.append(data[4][i])
            AC.append(data[5][i])

        # 모든 페이지를 크롤링했는지 체크
        if int(driver.find_elements_by_class_name("current")[0].text) == 1:
            break
        else: 
            # 현재 페이지 크롤링 종료시 이전 페이지로 이동
            beforeBtn = driver.find_elements_by_xpath('//*[@id="DataTables_Table_0_previous"]')
            beforeBtn[0].click()
    
    SN = len(PR)

    # 엑셀 파일 제어
    count = 2
    for i in range(SN):
        write_ws.cell(count, 1, i+1)
        write_ws.cell(count, 2, PR[i].text)
        write_ws.cell(count, 3, DT[i].text)
        write_ws.cell(count, 4, AR[i].text)
        write_ws.cell(count, 5, IF[i].text)
        write_ws.cell(count, 6, CT[i].text)
        write_ws.cell(count, 7, AC[i].text)
        count += 1
    write_wb.save(os.path.join(BASE_DIR, "result.xlsx"))
main()